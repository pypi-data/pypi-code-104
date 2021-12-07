import pandas as pd
from io import StringIO
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter


df = pd.read_csv(StringIO("""\
               alpha      beta     gamma
2000-01-01 -0.173215  0.119209 -1.044236
2000-01-02 -0.861849 -2.104569 -0.494929
2000-01-03  1.071804  0.721555 -0.706771
2000-01-04 -1.039575  0.271860 -0.424972
2000-01-05  0.567020  0.276232 -1.087401
2000-01-06 -0.673690  0.113648 -1.478427
2000-01-07  0.524988  0.404705  0.577046
2000-01-08 -1.715002 -1.039268 -0.370647
"""), sep="\s+", parse_dates=True)

output_filename = "salida.xlsx"


def no_customization(sheet_name='No customization'):
    df.to_excel(output_filename, sheet_name)


def little_customization(sheet_name='Little customization'):
    # Python context manager (with ... as ..:) automatically closes resource
    # when block is left, even under error conditions/cancellation/...
    with pd.ExcelWriter(
            output_filename,
            mode='a',  # append to existing file; default='w' (overwrite)
            engine='openpyxl') as xlsx:
        df.to_excel(xlsx, sheet_name)
        
        # set column width
        xlsx.sheets[sheet_name].column_dimensions['A'].width = 21


def full_customization(sheet_name='Full customization'):
    # Unpacking arguments using '**' for dictionary of keyword arguments
    # see: https://docs.python.org/3/tutorial/controlflow.html#more-on-defining-functions
    writer_args = {
        'path': output_filename,
        'mode': 'a',
        'engine': 'openpyxl'}
        
    # https://openpyxl.readthedocs.io/en/stable/formatting.html#colorscale
    percentile_rule = ColorScaleRule(
        start_type='percentile', start_value=10, start_color='ffaaaa',  # red-ish
        mid_type='num', mid_value=0, mid_color='ffffff',  # value zero==white
        end_type='percentile', end_value=90, end_color='aaffaa')  # green-ish
        
    # create a custom named style for the index
    index_style = NamedStyle(
        name="Index Style",
        number_format='YYYY-MM-DD, DDD',
        font=Font(color='999999', italic=True),
        alignment=Alignment(horizontal='left'))
    
    with pd.ExcelWriter(**writer_args) as xlsx:
        df.to_excel(xlsx, sheet_name)
        
        # worksheets that have been created with this ExcelWriter can be accessed
        # by openpyxl using its API. `ws` is now a openpyxl Worksheet object
        ws = xlsx.sheets[sheet_name]
        
        # cell ranges
        title_row = '1'
        value_cells = 'B2:{col}{row}'.format(
            col=get_column_letter(ws.max_column),
            row=ws.max_row)
        index_column = 'A'
        
        ws.column_dimensions[index_column].width = 21

        # which offers, among other things, a conditional_formatting facility
        ws.conditional_formatting.add(value_cells, percentile_rule)
        
        # for general styling, one has to iterate over all cells individually
        for row in ws[value_cells]:
            for cell in row:
                cell.number_format = '0.00'
        
        # builtin or named styles can be applied by using the object or their name
        # https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles
        for cell in ws[index_column]:
            cell.style = index_style
        
        # style header line last, so that headline style wins in cell A1
        for cell in ws[title_row]:
            cell.style = 'Headline 2'
            

no_customization()
little_customization()
full_customization()

#################################### 1.FUSION ######################################



########################################################################################
'''
#ID_P_T, X_t , y = get_df_procesado_temp(2019,'DESERCION_2019_2020',"6 prim",[9],"oriente",True,None,2019)
ID_P_T, X_t , y = get_df_procesado(2019,'DESERCION_2019_2020',"6 prim",[9],"oriente",True,None,2019)
test_size = hcm.get_test_size(X_t)
X_train, X_test, y_train, y_test= train_test_split(X_t, y, test_size=test_size,stratify=y,random_state=42)
model_name = "neg_bagging_fraction__lgb_model"
model , predicted_probas  , KPIs = hcm.modelar_clasificacion_binaria(model_name, X_train,y_train,X_test,y_test)
'''
#########################################################################################
'''
ID_P_T, X_t , y = get_df_procesado(2019,'DESERCION_2019_2020',"6 prim",[9],"oriente",True,None,2019)
X_t['randNumCol'] = np.random.normal(3, 2.5, X_t.shape[0])
print(hd.get_bool_columns(X_t))
inputs_num = list(set(X_t.columns.tolist())-set(hd.get_bool_columns(X_t)))


data_ca = {'Variable' : inputs_num,
       'Coeficiente_Asim' : stats.skew(X_t[inputs_num])}
asim = pd.DataFrame(data_ca)


inputs_num_sel = np.array(asim.iloc[:,0][asim['Coeficiente_Asim']<=5])
inputs_num_sel

res_data_num = X_t[inputs_num].describe().transpose()
res_data_num['cv'] = res_data_num.iloc[:,2] / res_data_num.iloc[:,1] * 100
res_data_num

from sklearn.preprocessing import RobustScaler
transformer = RobustScaler().fit(X_t)
data_inputs_s = transformer.transform(X_t)
X_t = pd.DataFrame(data_inputs_s,columns=X_t.columns.tolist())

sns.kdeplot(X_t['randNumCol'],shade = True,vertical = False,cumulative = False,color = "#BB0000")
plt.show()

test_size = hcm.get_test_size(X_t)
X_train, X_test, y_train, y_test= train_test_split(X_t, y, test_size=test_size,stratify=y,random_state=42)
       

model_name = "neg_bagging_fraction__lgb_model"
model , predicted_probas  , KPIs = hcm.modelar_clasificacion_binaria(model_name, X_train,y_train,X_test,y_test)


specific_url_model = "temp/model.sav"
model = pickle.load( open( specific_url_model, "rb" ) )

y_pred2 ,y_prob2_cb,  predicted_probas2 = hcm.predecir_clasificacion_binaria(model,X_test)

'''






  